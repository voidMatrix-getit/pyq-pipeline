"""
step4_xlsx.py
Converts final pipe-separated lines into a formatted Excel file.
Uses openpyxl (pip install openpyxl).
Admin-panel import compatible format.
"""

import logging
from pathlib import Path
from scripts.utils import parse_pipe_line, COLUMNS

log = logging.getLogger(__name__)

# Column widths in Excel units
COL_WIDTHS = {
    "No":             5,
    "Section":        22,
    "Sub-Section":    25,
    "Question":       55,
    "Answer 1":       28,
    "Answer 2":       28,
    "Answer 3":       28,
    "Answer 4":       28,
    "Correct Answer": 15,
    "Explanation":    55,
    "Difficulty":     12,
}

HEADER_FILL  = "1F4E79"
HEADER_FONT  = "FFFFFF"
ALT_FILL     = "EBF3FB"
UNCERTAIN_FG = "FF0000"


def generate_xlsx(lines: list[str], output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return

    rows = []
    for line in lines:
        row = parse_pipe_line(line)
        if row:
            rows.append(row)
        else:
            log.warning(f"Skipped malformed line: {line[:60]}")

    if not rows:
        log.error("No valid rows to write to XLSX.")
        return

    wb = Workbook()
    ws = wb.active
    assert ws is not None          # tells Pylance ws is never None
    ws.title = "Question Bank"

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font   = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
    hdr_fill   = PatternFill("solid", fgColor=HEADER_FILL)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font  = Font(name="Arial", size=9)
    body_align = Alignment(vertical="top", wrap_text=True)
    alt_fill   = PatternFill("solid", fgColor=ALT_FILL)

    unc_font   = Font(name="Arial", size=9, bold=True, color=UNCERTAIN_FG)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    for j, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    # ── Data rows ────────────────────────────────────────────────────────
    for i, row in enumerate(rows, 2):
        ws.row_dimensions[i].height = 80   # tall row for wrapped text
        fill = alt_fill if i % 2 == 0 else None
        for j, col in enumerate(COLUMNS, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border
            if col == "Correct Answer" and "UNCERTAIN" in val.upper():
                cell.font = unc_font
            else:
                cell.font = body_font
            cell.alignment = body_align
            if fill:
                cell.fill = fill

    # ── Column widths ────────────────────────────────────────────────────
    for j, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(col, 15)

    # ── Freeze panes & auto-filter ────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Second sheet: UNCERTAIN review list ──────────────────────────────
    uncertain_rows = [r for r in rows if "UNCERTAIN" in r.get("Correct Answer", "").upper()]
    if uncertain_rows:
        ws2 = wb.create_sheet("UNCERTAIN_Review")
        ws2.row_dimensions[1].height = 28
        for j, col in enumerate(COLUMNS, 1):
            cell = ws2.cell(row=1, column=j, value=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
        for i, row in enumerate(uncertain_rows, 2):
            for j, col in enumerate(COLUMNS, 1):
                val  = row.get(col, "")
                cell = ws2.cell(row=i, column=j, value=val)
                cell.font      = unc_font if col == "Correct Answer" else body_font
                cell.alignment = body_align
                cell.border    = border
        for j, col in enumerate(COLUMNS, 1):
            ws2.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(col, 15)
        log.info(f"  UNCERTAIN sheet: {len(uncertain_rows)} questions flagged for manual review")

    wb.save(output_path)
    log.info(f"XLSX written: {output_path} ({len(rows)} rows)")